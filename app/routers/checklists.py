"""Checklist template endpoints — engineer read + admin builder (CRUD)."""
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.deps import get_current_user, require_role
from app.models import ChecklistItem, ChecklistTemplate, User
from app.schemas import (
    ChecklistItemOut, ChecklistOut, ChecklistTemplateIn, ChecklistTemplateOut, TemplateSummary,
)

router = APIRouter(prefix="/checklists", tags=["checklists"])


@router.get("", response_model=ChecklistOut)
async def get_checklist(
    equipment_type: str,
    frequency: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    tpl = (await db.execute(
        select(ChecklistTemplate)
        .where(ChecklistTemplate.equipment_type == equipment_type, ChecklistTemplate.frequency == frequency)
        .options(selectinload(ChecklistTemplate.items))
    )).scalar_one_or_none()
    if tpl is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No checklist defined for {equipment_type} / {frequency}",
        )
    return ChecklistOut(
        equipment_type=tpl.equipment_type, frequency=tpl.frequency,
        items=[ChecklistItemOut.model_validate(i) for i in tpl.items],
    )


@router.get("/templates", response_model=list[TemplateSummary])
async def list_templates(
    equipment_type: str | None = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = select(ChecklistTemplate).options(selectinload(ChecklistTemplate.items))
    if equipment_type:
        q = q.where(ChecklistTemplate.equipment_type == equipment_type)
    rows = (await db.execute(q)).scalars().all()
    return [TemplateSummary(id=t.id, equipment_type=t.equipment_type, frequency=t.frequency, item_count=len(t.items)) for t in rows]


@router.get("/import-template")
async def checklist_import_template(_: User = Depends(require_role("admin"))):
    """Download a blank .xlsx template for bulk checklist import (one row per item)."""
    from io import BytesIO

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"
    ws.append(["equipment_type", "frequency", "section", "item_no", "description", "reading_code", "reading_unit", "requires_reading"])
    # A normal yes/no item
    ws.append(["MV Fan", "1M", "a) Vibration and noise", "i", "Loose or broken bolts", "", "", "no"])
    ws.append(["MV Fan", "1M", "a) Vibration and noise", "ii", "Visual check any abnormal vibration", "", "", "no"])
    # A 12M-only item (different frequency in the same sheet)
    ws.append(["MV Fan", "12M", "b) Fan casing and support", "iv", "Corrosion / metal fatigue of internal parts", "", "", "no"])
    # A single-reading item
    ws.append(["MV Fan", "1M", "c) Motor electrical properties", "iii", "Running voltage (Low Speed)", "VL", "V", ""])
    # A MULTI-reading item: three rows, same item, blank description on the 2nd/3rd -> readings IH1/IH2/IH3
    ws.append(["MV Fan", "1M", "c) Motor electrical properties", "ii", "Running current (High Speed)", "IH1", "A", ""])
    ws.append(["MV Fan", "1M", "c) Motor electrical properties", "ii", "", "IH2", "A", ""])
    ws.append(["MV Fan", "1M", "c) Motor electrical properties", "ii", "", "IH3", "A", ""])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=flowea_checklist_template.xlsx"},
    )


_CL_ALIASES = {"type": "equipment_type", "equipment type": "equipment_type", "freq": "frequency",
               "item": "item_no", "item no": "item_no", "no": "item_no", "sr.no": "item_no", "sr no": "item_no",
               "desc": "description", "maintenance description": "description",
               "section": "section", "group": "section", "category": "section",
               "code": "reading_code", "reading code": "reading_code", "parameter": "reading_code",
               "reading": "requires_reading", "needs reading": "requires_reading", "requires reading": "requires_reading",
               "unit": "reading_unit", "reading unit": "reading_unit"}


def _cl_header(h) -> str:
    h = str(h or "").strip().lower()
    return _CL_ALIASES.get(h, h.replace(" ", "_"))


def _truthy(v) -> bool:
    return str(v or "").strip().lower() in ("yes", "y", "true", "1", "x", "✓")


@router.post("/import")
async def import_checklists(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_role("admin")),
):
    """Bulk import checklist items from .xlsx/.csv. Rows are grouped by (equipment_type, frequency)."""
    raw = await file.read()
    fname = (file.filename or "").lower()
    records: list[dict] = []
    try:
        if fname.endswith(".csv"):
            import csv
            import io
            for row in csv.DictReader(io.StringIO(raw.decode("utf-8-sig"))):
                records.append({_cl_header(k): (v or "").strip() for k, v in row.items()})
        else:
            from io import BytesIO
            from openpyxl import load_workbook
            ws = load_workbook(BytesIO(raw), read_only=True, data_only=True).active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise ValueError("empty sheet")
            headers = [_cl_header(h) for h in rows[0]]
            for r in rows[1:]:
                if r is None or all(c is None or str(c).strip() == "" for c in r):
                    continue
                records.append({headers[i]: ("" if c is None else str(c).strip()) for i, c in enumerate(r) if i < len(headers)})
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read file: {e}")

    groups: dict[tuple, list] = {}
    errors = []
    last_item = None          # the most recent real item (for appending reading sub-rows)
    last_freq = None
    last_section = None
    for idx, rec in enumerate(records, start=2):
        etype = rec.get("equipment_type") or None
        freq = rec.get("frequency") or last_freq
        section = rec.get("section") or None
        desc = rec.get("description")
        code = rec.get("reading_code")
        unit = rec.get("reading_unit") or None

        # a sub-row: no description but a reading code -> add another reading to the previous item
        if not desc and code:
            if last_item is None:
                errors.append(f"Row {idx}: reading '{code}' has no item above it")
                continue
            last_item.readings = (last_item.readings or []) + [{"code": code, "unit": unit}]
            last_item.requires_reading = True
            continue

        if not desc:
            continue  # blank / spacer row
        if not etype or not freq:
            errors.append(f"Row {idx}: needs equipment_type and frequency")
            continue
        if section:
            last_section = section
        last_freq = freq
        items = groups.setdefault((etype, freq), [])
        item = ChecklistItem(
            item_no=rec.get("item_no") or str(len(items) + 1),
            section=section or last_section,
            description=desc,
            requires_reading=_truthy(rec.get("requires_reading")) or bool(code),
            reading_unit=unit if not code else None,
            readings=[{"code": code, "unit": unit}] if code else [],
            sort_order=len(items),
        )
        items.append(item)
        last_item = item

    saved = 0
    for (etype, freq), items in groups.items():
        existing = (await db.execute(
            select(ChecklistTemplate).where(ChecklistTemplate.equipment_type == etype, ChecklistTemplate.frequency == freq)
            .options(selectinload(ChecklistTemplate.items))
        )).scalar_one_or_none()
        if existing is None:
            existing = ChecklistTemplate(equipment_type=etype, frequency=freq)
            db.add(existing)
        existing.items = items
        saved += 1
    await db.commit()
    return {"checklists": saved, "items": sum(len(v) for v in groups.values()), "errors": errors}


@router.get("/templates/{template_id}", response_model=ChecklistTemplateOut)
async def get_template(template_id: uuid.UUID, db: AsyncSession = Depends(get_db), _: User = Depends(require_role("admin"))):
    t = (await db.execute(
        select(ChecklistTemplate).where(ChecklistTemplate.id == template_id).options(selectinload(ChecklistTemplate.items))
    )).scalar_one_or_none()
    if t is None:
        raise HTTPException(status_code=404, detail="Template not found")
    return ChecklistTemplateOut(id=t.id, equipment_type=t.equipment_type, frequency=t.frequency,
                                items=[ChecklistItemOut.model_validate(i) for i in t.items])


async def _upsert(db, body: ChecklistTemplateIn) -> ChecklistTemplate:
    t = (await db.execute(
        select(ChecklistTemplate)
        .where(ChecklistTemplate.equipment_type == body.equipment_type, ChecklistTemplate.frequency == body.frequency)
        .options(selectinload(ChecklistTemplate.items))
    )).scalar_one_or_none()
    if t is None:
        t = ChecklistTemplate(equipment_type=body.equipment_type, frequency=body.frequency)
        db.add(t)
    t.items = [
        ChecklistItem(
            item_no=i.item_no, section=i.section, description=i.description,
            requires_reading=i.requires_reading or bool(i.readings),
            reading_unit=i.reading_unit,
            readings=[r.model_dump() for r in i.readings] if i.readings else [],
            sort_order=idx,
        )
        for idx, i in enumerate(body.items)
    ]
    return t


@router.put("/templates", response_model=ChecklistTemplateOut)
async def save_template(
    body: ChecklistTemplateIn,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_role("admin")),
):
    """Create or replace a checklist template (keyed by equipment type + frequency)."""
    if not body.items:
        raise HTTPException(status_code=422, detail="A checklist needs at least one item")
    t = await _upsert(db, body)
    await db.commit()
    await db.refresh(t, attribute_names=["items"])
    return ChecklistTemplateOut(id=t.id, equipment_type=t.equipment_type, frequency=t.frequency,
                                items=[ChecklistItemOut.model_validate(i) for i in t.items])


@router.delete("/templates/{template_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_template(template_id: uuid.UUID, db: AsyncSession = Depends(get_db), _: User = Depends(require_role("admin"))):
    t = await db.get(ChecklistTemplate, template_id)
    if t:
        await db.delete(t)
        await db.commit()
