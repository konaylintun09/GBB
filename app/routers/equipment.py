"""Equipment + schedule endpoints."""
import uuid
from datetime import date

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.deps import get_current_user, require_role
from app.models import Equipment, MaintenanceSchedule, User
from app.schemas import EquipmentCreate, EquipmentOut, EquipmentUpdate, MetaOut, ScheduleOut
from app.util import days_until, next_due_from, parse_frequency, schedule_status

router = APIRouter(prefix="/equipment", tags=["equipment"])


def _serialize(eq: Equipment) -> EquipmentOut:
    sched = None
    if eq.schedule:
        sched = ScheduleOut(
            frequency=eq.schedule.frequency,
            interval_value=eq.schedule.interval_value,
            interval_unit=eq.schedule.interval_unit,
            last_done=eq.schedule.last_done,
            next_due=eq.schedule.next_due,
            status=schedule_status(eq.schedule.next_due),
            days_until=days_until(eq.schedule.next_due),
        )
    return EquipmentOut(
        id=eq.id, tag=eq.tag, name=eq.name, type=eq.type, location=eq.location,
        site_name=eq.site_name, system=eq.system, status=eq.status,
        cooling_capacity=eq.cooling_capacity, unit_type=eq.unit_type, qr_token=eq.qr_token, schedule=sched,
    )


def _resolve_interval(body) -> tuple[str, int, str]:
    """Return (frequency_label, interval_value, interval_unit) from a create/update body."""
    if body.interval_value and body.interval_unit:
        value, unit = body.interval_value, body.interval_unit
        label = body.frequency or f"{value}{unit[0].upper()}"
    else:
        value, unit = parse_frequency(body.frequency or "1M")
        label = body.frequency or f"{value}{unit[0].upper()}"
    return label, value, unit


@router.get("", response_model=list[EquipmentOut])
async def list_equipment(db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    rows = (await db.execute(select(Equipment).options(selectinload(Equipment.schedule)))).scalars().all()
    return [_serialize(e) for e in rows]


@router.get("/meta", response_model=MetaOut)
async def meta(db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    rows = (await db.execute(select(Equipment).options(selectinload(Equipment.schedule)))).scalars().all()
    types = sorted({e.type for e in rows if e.type})
    systems = sorted({e.system for e in rows if e.system})
    units = sorted({e.unit_type for e in rows if e.unit_type})
    sites = sorted({e.site_name for e in rows if e.site_name})
    freqs = sorted({e.schedule.frequency for e in rows if e.schedule})
    # always offer the common defaults as suggestions
    for d in ("ACMV", "Electrical", "Fire", "Plumbing"):
        if d not in systems:
            systems.append(d)
    for f in ("1M", "3M", "6M", "12M"):
        if f not in freqs:
            freqs.append(f)
    return MetaOut(types=types, systems=systems, unit_types=units or ["CHW", "DX"], frequencies=freqs, sites=sites)


@router.get("/by-qr/{qr_token}", response_model=EquipmentOut)
async def scan(qr_token: str, db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    eq = (await db.execute(
        select(Equipment).where(Equipment.qr_token == qr_token).options(selectinload(Equipment.schedule))
    )).scalar_one_or_none()
    if eq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Unknown or revoked QR code")
    return _serialize(eq)


# ---- Excel / CSV bulk import ----
_IMPORT_COLUMNS = ["name", "tag", "type", "location", "site_name", "system", "cooling_capacity", "unit_type", "frequency"]
_ALIASES = {"site": "site_name", "capacity": "cooling_capacity", "cooling capacity": "cooling_capacity",
            "unit type": "unit_type", "freq": "frequency", "equipment name": "name", "equipment tag": "tag"}


def _norm_header(h) -> str:
    h = str(h or "").strip().lower()
    return _ALIASES.get(h, h.replace(" ", "_"))


@router.get("/import-template")
async def import_template(_: User = Depends(require_role("admin"))):
    """Download a blank .xlsx template with the expected columns."""
    from io import BytesIO

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment"
    ws.append(["name", "tag", "type", "location", "site_name", "system", "cooling_capacity", "unit_type", "frequency"])
    ws.append(["Air Handling Unit 9", "", "AHU", "Level 9", "Marina Bay Tower", "ACMV", "18000 CMH", "CHW", "3M"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=flowea_equipment_template.xlsx"},
    )


@router.post("/import")
async def import_equipment(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_role("admin")),
):
    """Bulk-create equipment from an uploaded .xlsx or .csv file."""
    raw = await file.read()
    fname = (file.filename or "").lower()
    records: list[dict] = []
    try:
        if fname.endswith(".csv"):
            import csv
            import io
            reader = csv.DictReader(io.StringIO(raw.decode("utf-8-sig")))
            for row in reader:
                records.append({_norm_header(k): (v or "").strip() for k, v in row.items()})
        else:
            from io import BytesIO
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise ValueError("empty sheet")
            headers = [_norm_header(h) for h in rows[0]]
            for r in rows[1:]:
                if r is None or all(c is None or str(c).strip() == "" for c in r):
                    continue
                records.append({headers[i]: ("" if c is None else str(c).strip()) for i, c in enumerate(r) if i < len(headers)})
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read file: {e}")

    created, errors = 0, []
    for idx, rec in enumerate(records, start=2):
        try:
            body = EquipmentCreate(
                name=rec.get("name") or None, tag=rec.get("tag") or None, type=rec.get("type") or None,
                location=rec.get("location") or None, site_name=rec.get("site_name") or None,
                system=rec.get("system") or None, cooling_capacity=rec.get("cooling_capacity") or None,
                unit_type=rec.get("unit_type") or None, frequency=rec.get("frequency") or None,
            )
            await _build_equipment(db, body)
            created += 1
        except Exception as e:
            errors.append(f"Row {idx}: {e}")
    await db.commit()
    return {"created": created, "errors": errors}


@router.get("/{equipment_id}", response_model=EquipmentOut)
async def get_one(equipment_id: uuid.UUID, db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    eq = (await db.execute(
        select(Equipment).where(Equipment.id == equipment_id).options(selectinload(Equipment.schedule))
    )).scalar_one_or_none()
    if eq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Equipment not found")
    return _serialize(eq)


async def _build_equipment(db: AsyncSession, body) -> Equipment:
    etype = (body.type or "General").strip()
    count = len((await db.execute(select(Equipment.id))).scalars().all())
    tag = (body.tag or "").strip() or f"{etype[:3].upper()}-{count + 1:04d}"
    if (await db.execute(select(Equipment.id).where(Equipment.tag == tag))).first():
        tag = f"{tag}-{uuid.uuid4().hex[:4]}"
    label, value, unit = _resolve_interval(body)
    eq = Equipment(
        tag=tag,
        name=(body.name or "Untitled equipment").strip(),
        type=etype,
        location=(body.location or "").strip(),
        site_name=(body.site_name or None),
        system=(body.system or "").strip(),
        status=body.status or "Active",
        cooling_capacity=body.cooling_capacity,
        unit_type=body.unit_type,
        qr_token=f"FLW-{uuid.uuid4().hex[:16]}",
    )
    eq.schedule = MaintenanceSchedule(
        frequency=label, interval_value=value, interval_unit=unit,
        last_done=None, next_due=next_due_from(date.today(), value, unit),
    )
    db.add(eq)
    return eq


@router.post("", response_model=EquipmentOut, status_code=status.HTTP_201_CREATED)
async def create_equipment(
    body: EquipmentCreate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_role("admin")),
):
    eq = await _build_equipment(db, body)
    await db.commit()
    await db.refresh(eq, attribute_names=["schedule"])
    return _serialize(eq)


@router.patch("/{equipment_id}", response_model=EquipmentOut)
async def update_equipment(
    equipment_id: uuid.UUID,
    body: EquipmentUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_role("admin")),
):
    eq = (await db.execute(
        select(Equipment).where(Equipment.id == equipment_id).options(selectinload(Equipment.schedule))
    )).scalar_one_or_none()
    if eq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Equipment not found")

    for f in ("name", "type", "location", "site_name", "system", "cooling_capacity", "unit_type", "status"):
        v = getattr(body, f)
        if v is not None:
            setattr(eq, f, v)

    if body.frequency or (body.interval_value and body.interval_unit):
        label, value, unit = _resolve_interval(body)
        eq.schedule.frequency = label
        eq.schedule.interval_value = value
        eq.schedule.interval_unit = unit
        base = eq.schedule.last_done or date.today()
        eq.schedule.next_due = next_due_from(base, value, unit)

    await db.commit()
    await db.refresh(eq, attribute_names=["schedule"])
    return _serialize(eq)
